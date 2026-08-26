from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from io import BytesIO
import html
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

app = FastAPI(title="Retenciones emitidas SRI")

MAX_UPLOAD_MB = 50
MAX_XML_FILES = 3000
MAX_UNCOMPRESSED_MB = 150

HTML = r"""
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Retenciones emitidas SRI</title>
<style>
:root{--blue:#0f4c81;--line:#d9dee7;--soft:#f4f7fb;--text:#172033;--ok:#0d7a45;--bad:#a61b1b}
*{box-sizing:border-box} body{font-family:Arial,sans-serif;max-width:1250px;margin:32px auto;padding:0 18px;color:var(--text)}
h1{margin-bottom:8px}.sub{margin-top:0;color:#4b5563}.card{border:1px solid var(--line);border-radius:14px;padding:18px;margin:18px 0}
.controls{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
button{background:var(--blue);color:white;border:0;border-radius:8px;padding:11px 16px;cursor:pointer;font-weight:600}
button.secondary{background:#6b89a5} button:disabled{opacity:.5;cursor:not-allowed}
input[type=file]{max-width:100%;padding:8px}
.badge{display:inline-block;background:#eef2ff;padding:7px 10px;border-radius:999px}
#status{white-space:pre-wrap;margin-top:12px;color:#374151}
.stats{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:16px 0}
.stat{background:var(--soft);border:1px solid var(--line);border-radius:12px;padding:14px}.stat small{display:block;color:#64748b;margin-bottom:5px}.stat strong{font-size:20px}
.tablewrap{overflow:auto;border:1px solid var(--line);border-radius:10px}
table{border-collapse:collapse;width:100%;font-size:13px;min-width:1100px}
th,td{border-bottom:1px solid var(--line);padding:8px;text-align:left;vertical-align:top}th{background:#f3f4f6;position:sticky;top:0}
.num{text-align:right;white-space:nowrap}.ok{color:var(--ok)}.bad{color:var(--bad)}
@media(max-width:760px){.stats{grid-template-columns:repeat(2,minmax(0,1fr))}}
</style>
</head>
<body>
<h1>Retenciones emitidas SRI</h1>
<p class="sub">Carga el ZIP descargado del SRI. La aplicación lee los XML incluidos en el ZIP, ignora los PDF y consolida bases y valores retenidos. No necesita usuario ni contraseña del SRI.</p>

<div class="card">
  <div class="controls">
    <input id="file" type="file" accept=".zip,.xml">
    <button id="process">Procesar archivo</button>
    <button id="excel" class="secondary" disabled>Descargar Excel</button>
    <span id="count" class="badge">0 registros</span>
  </div>
  <div id="status" aria-live="polite">Selecciona un ZIP del SRI.</div>
</div>

<div class="stats">
  <div class="stat"><small>XML procesados</small><strong id="xmlCount">0</strong></div>
  <div class="stat"><small>PDF ignorados</small><strong id="pdfCount">0</strong></div>
  <div class="stat"><small>Total Renta retenida</small><strong id="rentaTotal">$0.00</strong></div>
  <div class="stat"><small>Total IVA retenido</small><strong id="ivaTotal">$0.00</strong></div>
</div>

<div class="tablewrap">
<table id="tbl">
<thead><tr>
<th>Fecha</th><th>Proveedor</th><th>RUC</th><th>Retención</th><th>Factura</th>
<th>Impuesto</th><th>Código</th><th class="num">Base</th><th class="num">%</th>
<th class="num">Valor retenido</th><th>Estado</th>
</tr></thead>
<tbody></tbody>
</table>
</div>

<script>
let lastRows=[];
const $=id=>document.getElementById(id);
const f=$('file'), p=$('process'), e=$('excel'), s=$('status'), b=document.querySelector('#tbl tbody'), c=$('count');

function money(v){
  const n=Number(v||0);
  return n.toLocaleString('es-EC',{minimumFractionDigits:2,maximumFractionDigits:2});
}
function td(text, cls=''){
  const cell=document.createElement('td');
  cell.textContent=text ?? '';
  if(cls) cell.className=cls;
  return cell;
}
function render(rows){
  b.textContent='';
  const frag=document.createDocumentFragment();
  for(const x of rows){
    const tr=document.createElement('tr');
    tr.append(
      td(x.fecha),td(x.proveedor),td(x.ruc),td(x.retencion),td(x.factura),
      td(x.impuesto),td(x.codigo),
      td(money(x.base),'num'),td(money(x.porcentaje),'num'),
      td(money(x.valor_retenido),'num'),td(x.estado)
    );
    frag.appendChild(tr);
  }
  b.appendChild(frag);
}
p.addEventListener('click', async()=>{
  if(!f.files.length){s.textContent='Selecciona un archivo ZIP o XML del SRI.';return;}
  p.disabled=true; e.disabled=true; b.textContent=''; lastRows=[];
  s.textContent='Procesando XML del SRI...';
  const fd=new FormData(); fd.append('file',f.files[0]);
  try{
    const r=await fetch('/process',{method:'POST',body:fd});
    const raw=await r.text();
    let data;
    try{ data=JSON.parse(raw); }catch{ throw new Error(raw || 'El servidor devolvió una respuesta no válida.'); }
    if(!r.ok) throw new Error(data.detail || 'Error al procesar el archivo.');
    lastRows=data.rows || [];
    render(lastRows);
    c.textContent=lastRows.length+' registros';
    $('xmlCount').textContent=data.xml_processed || 0;
    $('pdfCount').textContent=data.pdf_ignored || 0;
    $('rentaTotal').textContent='$'+money(data.summary?.renta || 0);
    $('ivaTotal').textContent='$'+money(data.summary?.iva || 0);
    const err=data.xml_errors || 0;
    s.textContent=`Listo. XML procesados: ${data.xml_processed}. PDF ignorados: ${data.pdf_ignored}. XML con error: ${err}.`;
    e.disabled=lastRows.length===0;
  }catch(err){
    s.textContent='Error: '+err.message;
  }finally{p.disabled=false;}
});
e.addEventListener('click', async()=>{
  if(!lastRows.length) return;
  e.disabled=true;
  try{
    const r=await fetch('/excel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rows:lastRows})});
    if(!r.ok) throw new Error('No se pudo generar el Excel.');
    const blob=await r.blob();
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob); a.download='retenciones_sri.xlsx';
    document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(a.href);
  }catch(err){s.textContent='Error: '+err.message;} finally{e.disabled=false;}
});
</script>
</body>
</html>
"""

def lname(tag):
    return tag.split("}")[-1] if "}" in tag else tag

def child_text(node, name, default=""):
    for ch in list(node):
        if lname(ch.tag) == name:
            return (ch.text or "").strip()
    return default

def first_text(root, name, default=""):
    for el in root.iter():
        if lname(el.tag) == name:
            return (el.text or "").strip()
    return default

def descendants(root, name):
    return [el for el in root.iter() if lname(el.tag) == name]

def to_float(value):
    try:
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return 0.0

def unwrap_comprobante(root):
    # Acepta tanto XML directo de comprobanteRetencion como XML de autorización
    if lname(root.tag) == "comprobanteRetencion":
        return root
    for el in root.iter():
        if lname(el.tag) == "comprobante":
            txt = (el.text or "").strip()
            if txt.startswith("<"):
                return ET.fromstring(txt)
    # Algunos contenedores pueden incluir el comprobante como nodo hijo real.
    for el in root.iter():
        if lname(el.tag) == "comprobanteRetencion":
            return el
    return root

def parse_retention_xml(xml_bytes, source_name=""):
    root = ET.fromstring(xml_bytes)
    root = unwrap_comprobante(root)

    proveedor = first_text(root, "razonSocialSujetoRetenido")
    ruc = first_text(root, "identificacionSujetoRetenido")
    fecha = first_text(root, "fechaEmision")
    estab = first_text(root, "estab")
    pto = first_text(root, "ptoEmi")
    sec = first_text(root, "secuencial")
    retencion = "-".join(x for x in (estab, pto, sec) if x)
    if not retencion:
        retencion = first_text(root, "claveAcceso") or source_name

    estado = "PROCESADO"
    rows = []

    # Esquema 2.x: docSustento -> retenciones -> retencion
    doc_nodes = descendants(root, "docSustento")
    for doc in doc_nodes:
        factura = child_text(doc, "numDocSustento") or first_text(doc, "numDocSustento")
        for ret in descendants(doc, "retencion"):
            codigo_impuesto = child_text(ret, "codigo")
            codigo_ret = child_text(ret, "codigoRetencion")
            base = child_text(ret, "baseImponible")
            porcentaje = child_text(ret, "porcentajeRetener")
            valor = child_text(ret, "valorRetenido")
            if not any([base, porcentaje, valor, codigo_ret]):
                continue
            impuesto = {"1":"RENTA","2":"IVA","6":"ISD"}.get(codigo_impuesto, codigo_impuesto or "")
            rows.append({
                "fecha": fecha, "proveedor": proveedor, "ruc": ruc, "retencion": retencion,
                "factura": factura, "impuesto": impuesto, "codigo": codigo_ret or codigo_impuesto,
                "base": to_float(base), "porcentaje": to_float(porcentaje),
                "valor_retenido": to_float(valor), "estado": estado
            })

    if rows:
        return rows

    # Esquema 1.0.0: impuestos -> impuesto
    for imp in descendants(root, "impuesto"):
        codigo_impuesto = child_text(imp, "codigo")
        codigo_ret = child_text(imp, "codigoRetencion")
        base = child_text(imp, "baseImponible")
        porcentaje = child_text(imp, "porcentajeRetener")
        valor = child_text(imp, "valorRetenido")
        factura = child_text(imp, "numDocSustento") or first_text(root, "numDocSustento")
        if not any([base, porcentaje, valor]):
            continue
        impuesto = {"1":"RENTA","2":"IVA","6":"ISD"}.get(codigo_impuesto, codigo_impuesto or "")
        rows.append({
            "fecha": fecha, "proveedor": proveedor, "ruc": ruc, "retencion": retencion,
            "factura": factura, "impuesto": impuesto, "codigo": codigo_ret or codigo_impuesto,
            "base": to_float(base), "porcentaje": to_float(porcentaje),
            "valor_retenido": to_float(valor), "estado": estado
        })
    return rows

@app.get("/", response_class=HTMLResponse)
def home():
    return HTML

@app.post("/process")
async def process(file: UploadFile = File(...)):
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_MB * 1024 * 1024:
        return JSONResponse({"detail": f"El archivo supera el límite de {MAX_UPLOAD_MB} MB."}, status_code=413)

    filename = (file.filename or "").lower()
    rows = []
    xml_processed = 0
    pdf_ignored = 0
    xml_errors = 0

    if filename.endswith(".xml"):
        try:
            rows.extend(parse_retention_xml(raw, file.filename or "archivo.xml"))
            xml_processed = 1
        except Exception as exc:
            return JSONResponse({"detail": f"No pude leer el XML: {type(exc).__name__}."}, status_code=400)

    elif filename.endswith(".zip"):
        try:
            zf = zipfile.ZipFile(BytesIO(raw))
        except zipfile.BadZipFile:
            return JSONResponse({"detail": "El archivo no es un ZIP válido."}, status_code=400)

        infos = [i for i in zf.infolist() if not i.is_dir()]
        xml_infos = [i for i in infos if i.filename.lower().endswith(".xml")]
        pdf_ignored = sum(1 for i in infos if i.filename.lower().endswith(".pdf"))

        if not xml_infos:
            return JSONResponse({"detail": "No encontré archivos XML dentro del ZIP."}, status_code=400)
        if len(xml_infos) > MAX_XML_FILES:
            return JSONResponse({"detail": f"El ZIP contiene más de {MAX_XML_FILES} XML; divídelo en varios archivos."}, status_code=400)

        total_uncompressed = sum(i.file_size for i in xml_infos)
        if total_uncompressed > MAX_UNCOMPRESSED_MB * 1024 * 1024:
            return JSONResponse({"detail": "El contenido XML descomprimido es demasiado grande."}, status_code=413)

        for info in xml_infos:
            try:
                xml_bytes = zf.read(info)
                parsed = parse_retention_xml(xml_bytes, info.filename)
                if parsed:
                    rows.extend(parsed)
                xml_processed += 1
            except Exception:
                xml_errors += 1
    else:
        return JSONResponse({"detail": "Formato no admitido. Sube el ZIP descargado del SRI o un XML."}, status_code=400)

    if not rows:
        return JSONResponse({
            "detail": "Se leyeron los XML, pero no encontré detalles de retención. Verifica que el ZIP corresponda a comprobantes de retención."
        }, status_code=400)

    rows.sort(key=lambda x: (x.get("fecha",""), x.get("retencion",""), x.get("factura","")))

    renta = sum(x["valor_retenido"] for x in rows if x["impuesto"] == "RENTA")
    iva = sum(x["valor_retenido"] for x in rows if x["impuesto"] == "IVA")
    total = sum(x["valor_retenido"] for x in rows)

    return {
        "rows": rows,
        "xml_processed": xml_processed,
        "pdf_ignored": pdf_ignored,
        "xml_errors": xml_errors,
        "summary": {"renta": round(renta,2), "iva": round(iva,2), "total": round(total,2)}
    }

@app.post("/excel")
async def excel(payload: dict):
    rows = payload.get("rows") or []
    if not rows:
        return JSONResponse({"detail":"No hay registros para exportar."}, status_code=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Retenciones"

    headers = ["Fecha","Proveedor","RUC","Retención","Factura","Impuesto","Código","Base","%","Valor retenido","Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for x in rows:
        ws.append([
            x.get("fecha",""), x.get("proveedor",""), x.get("ruc",""), x.get("retencion",""),
            x.get("factura",""), x.get("impuesto",""), x.get("codigo",""),
            to_float(x.get("base",0)), to_float(x.get("porcentaje",0)),
            to_float(x.get("valor_retenido",0)), x.get("estado","")
        ])

    for row in ws.iter_rows(min_row=2):
        row[7].number_format = '#,##0.00'
        row[8].number_format = '0.00'
        row[9].number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = [13,38,16,22,22,12,14,14,10,18,14]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Hoja resumen
    resumen = wb.create_sheet("Resumen")
    resumen.append(["Concepto","Valor"])
    resumen["A1"].font = resumen["B1"].font = Font(bold=True)
    renta = sum(to_float(x.get("valor_retenido",0)) for x in rows if x.get("impuesto") == "RENTA")
    iva = sum(to_float(x.get("valor_retenido",0)) for x in rows if x.get("impuesto") == "IVA")
    total = sum(to_float(x.get("valor_retenido",0)) for x in rows)
    resumen.append(["Registros", len(rows)])
    resumen.append(["Total Renta retenida", renta])
    resumen.append(["Total IVA retenido", iva])
    resumen.append(["Total retenido", total])
    for cell in ("B3","B4","B5"):
        resumen[cell].number_format = '#,##0.00'
    resumen.column_dimensions["A"].width = 28
    resumen.column_dimensions["B"].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=retenciones_sri.xlsx"}
    )
